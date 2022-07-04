import openpyxl
from openpyxl.styles import PatternFill


wb = openpyxl.load_workbook("book1.xlsx") #path to the Excel file
ws = wb.active #Name of the working sheet


fill_cell = PatternFill(patternType='solid', 
                            fgColor='C64747') #You can give the hex code for different color


# for row in ws.max_row:
#     print(row)
for row in ws.iter_rows():
    # print(row)
    cell1, cell2 = row
    # print(cell1)
    cell1.fill = fill_cell
    cell2.fill = fill_cell
wb.save("book1.xlsx")    
# for row in ws.iter_rows('C{}:C{}'.format(ws.min_row,ws.max_row)):
#     for cell in row:
#         print (cell.value)

# for i in ws.row_dimensions:
#     print

# ws[1][1].fill = fill_cell # B2 is the name of the cell to be fill with color
# wb.save("book1.xlsx")